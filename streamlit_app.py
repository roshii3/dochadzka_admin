import streamlit as st
import pandas as pd
from datetime import datetime, date, time, timedelta
import pytz
from supabase import create_client, Client
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

import streamlit as st

st._config.set_option("server.enableWatchdog", False)
st._config.set_option("server.runOnSave", False)


st.set_page_config(page_title="Admin - Dochádzka", layout="wide", initial_sidebar_state="expanded")
st.markdown("<style>#MainMenu{visibility:hidden;}footer{visibility:hidden;}header{visibility:hidden;}</style>", unsafe_allow_html=True)

DATABAZA_URL = st.secrets["DATABAZA_URL"]
DATABAZA_KEY = st.secrets["DATABAZA_KEY"]
ADMIN_PASS = st.secrets.get("ADMIN_PASS", "")
databaze: Client = create_client(DATABAZA_URL, DATABAZA_KEY)
tz = pytz.timezone("Europe/Bratislava")

POSITIONS = ["Veliteľ", "CCTV", "Brány", "Sklad2", "Sklad3", "Turniket2", "Turniket3", "Plombovac2", "Plombovac3"]
SHIFT_HOURS = 7.5
DOUBLE_SHIFT_HOURS = 15.25
VELITEL_DOUBLE = 16.25
SWAP_WINDOW_MINUTES = 30

def load_attendance(start_dt: datetime, end_dt: datetime) -> pd.DataFrame:
    res = databaze.table("attendance").select("*").gte("timestamp", start_dt.isoformat()).lt("timestamp", end_dt.isoformat()).execute()
    df = pd.DataFrame(res.data)
    if df.empty:
        return df
    df["timestamp"] = pd.to_datetime(df["timestamp"], errors="coerce")
    df["timestamp"] = df["timestamp"].apply(lambda x: tz.localize(x) if pd.notna(x) and x.tzinfo is None else x)
    def assign_logical_date(ts):
        if ts.time() < time(3,0):
            return (ts - timedelta(days=1)).date()
        return ts.date()
    df["date"] = df["timestamp"].apply(assign_logical_date)
    df["time"] = df["timestamp"].dt.time
    return df

def get_user_pairs(pos_day_df: pd.DataFrame):
    pairs = {}
    if pos_day_df.empty:
        return pairs
    for user in pos_day_df["user_code"].unique():
        u = pos_day_df[pos_day_df["user_code"] == user]
        pr = u[u["action"].str.lower() == "príchod"]["timestamp"]
        od = u[u["action"].str.lower() == "odchod"]["timestamp"]
        pr_min = pr.min() if not pr.empty else pd.NaT
        od_max = od.max() if not od.empty else pd.NaT
        pairs[user] = {"pr": pr_min, "od": od_max, "pr_count": len(pr), "od_count": len(od)}
    return pairs

def classify_pair(pr, od, position):
    msgs=[]
    if (pd.isna(pr) or pr is None) and (pd.isna(od) or od is None):
        return ("none","none",0.0,0.0,msgs)
    if pd.isna(pr) or pr is None:
        msgs.append("missing_prichod")
        return ("missing_pr","none",0.0,0.0,msgs)
    if pd.isna(od) or od is None:
        msgs.append("missing_odchod")
        return ("none","missing_od",0.0,0.0,msgs)
    pr_t = pr.time()
    od_t = od.time()
    if position.lower().startswith("vel"):
        if pr_t <= time(7,0) and (od_t >= time(21,0) or od_t < time(2,0)):
            return ("R+P OK","R+P OK",VELITEL_DOUBLE,VELITEL_DOUBLE,msgs)
    if pr_t <= time(7,0) and (od_t >= time(21,0) or od_t < time(2,0)):
        return ("R+P OK","R+P OK",DOUBLE_SHIFT_HOURS,DOUBLE_SHIFT_HOURS,msgs)
    if pr_t <= time(7,0) and od_t <= time(15,0):
        return ("Ranna OK","none",SHIFT_HOURS,0.0,msgs)
    if pr_t >= time(13,0) and od_t >= time(21,0):
        return ("none","Poobedna OK",0.0,SHIFT_HOURS,msgs)
    msgs.append("invalid_times")
    return ("invalid","invalid",0.0,0.0,msgs)

def merge_intervals(pairs):
    intervals=[]
    for pair in pairs.values():
        if pd.notna(pair["pr"]) and pd.notna(pair["od"]):
            intervals.append((pair["pr"],pair["od"]))
    if not intervals:
        return []
    intervals.sort(key=lambda x:x[0])
    merged=[intervals[0]]
    for start,end in intervals[1:]:
        last_start,last_end=merged[-1]
        gap_min=(start-last_end).total_seconds()/60
        if gap_min <= SWAP_WINDOW_MINUTES:
            merged[-1]=(last_start,max(last_end,end))
        else:
            merged.append((start,end))
    return merged

def summarize_position_day(pos_day_df: pd.DataFrame, position):
    morning={"status":"absent","hours":0.0,"detail":None}
    afternoon={"status":"absent","hours":0.0,"detail":None}
    details=[]
    if pos_day_df.empty:
        return morning,afternoon,details
    pairs=get_user_pairs(pos_day_df)
    rp_user=None
    for user,pair in pairs.items():
        role_m,role_p,h_m,h_p,msgs=classify_pair(pair["pr"],pair["od"],position)
        if role_m=="R+P OK" and role_p=="R+P OK":
            rp_user=(user,pair,h_m,h_p)
            break
    if rp_user:
        user,pair,h_m,h_p=rp_user
        morning={"status":"R+P OK","hours":h_m,"detail":f"Príchod: {pair['pr']}, Odchod: {pair['od']}"}
        afternoon={"status":"R+P OK","hours":h_p,"detail":f"Príchod: {pair['pr']}, Odchod: {pair['od']}"}
        return morning,afternoon,details
    had_invalid_or_missing=False
    for user,pair in pairs.items():
        role_m,role_p,h_m,h_p,msgs=classify_pair(pair["pr"],pair["od"],position)
        if role_m=="Ranna OK" and morning["status"] not in ("Ranna OK","R+P OK"):
            morning={"status":"Ranna OK","hours":h_m,"detail":f"{user}: Príchod: {pair['pr']}, Odchod: {pair['od']}"}
        if role_p=="Poobedna OK" and afternoon["status"] not in ("Poobedna OK","R+P OK"):
            afternoon={"status":"Poobedna OK","hours":h_p,"detail":f"{user}: Príchod: {pair['pr']}, Odchod: {pair['od']}"}
        if msgs:
            had_invalid_or_missing=True
            for m in msgs:
                details.append(f"{user}: {m} — pr:{pair['pr']} od:{pair['od']}")
    merged=merge_intervals(pairs)
    total_hours=round(sum((end-start).total_seconds()/3600 for start,end in merged),2) if merged else 0.0
    if merged:
        earliest=min(s[0] for s in merged)
        latest=max(s[1] for s in merged)
        e_t=earliest.time()
        l_t=latest.time()
        double_threshold=VELITEL_DOUBLE if position.lower().startswith("vel") else DOUBLE_SHIFT_HOURS
        if e_t <= time(7,0) and (l_t >= time(21,0) or l_t < time(2,0)) and total_hours >= double_threshold-0.01:
            morning={"status":"R+P OK","hours":round(total_hours/2,2),"detail":" + ".join([f"{u}: {p['pr']}–{p['od']}" for u,p in pairs.items()])}
            afternoon={"status":"R+P OK","hours":round(total_hours/2,2),"detail":morning["detail"]}
            return morning,afternoon,details
        if e_t <= time(7,0) and latest.time() <= time(15,0) and total_hours >= SHIFT_HOURS-0.01:
            morning={"status":"Ranna OK","hours":round(total_hours,2),"detail":" + ".join([f"{u}: {p['pr']}–{p['od']}" for u,p in pairs.items()])}
            return morning,afternoon,details
        if e_t >= time(13,0) and l_t >= time(21,0) and total_hours >= SHIFT_HOURS-0.01:
            afternoon={"status":"Poobedna OK","hours":round(total_hours,2),"detail":" + ".join([f"{u}: {p['pr']}–{p['od']}" for u,p in pairs.items()])}
            return morning,afternoon,details
    return morning,afternoon,details

def summarize_day(df_day: pd.DataFrame, target_date: date):
    results={}
    extra_velitel=False
    extra_sbs=False
    for pos in POSITIONS:
        pos_df=df_day[df_day["position"]==pos]
        morning,afternoon,details=summarize_position_day(pos_df,pos)
        for _,row in pos_df.iterrows():
            if row["action"].lower()=="odchod" and row["timestamp"].time()<time(3,0):
                if pos.lower().startswith("vel"):
                    extra_velitel=True
                else:
                    extra_sbs=True
        total=VELITEL_DOUBLE if pos.lower().startswith("vel") else DOUBLE_SHIFT_HOURS
        results[pos]={"morning":morning,"afternoon":afternoon,"details":details,"total_hours":total}
    if extra_velitel:
        results["Extra1 – Veliteľ"]={"morning":{"status":"EXTRA","hours":3.45},"afternoon":{"status":"","hours":0},"details":[],"total_hours":3.45}
    if extra_sbs:
        results["Extra2 – SBS"]={"morning":{"status":"EXTRA","hours":3.45},"afternoon":{"status":"","hours":0},"details":[],"total_hours":3.45}
    return results

def save_attendance(user_code, position, action, now=None):
    user_code=user_code.strip()
    if not now:
        now=datetime.now(tz)
    if now.second==0 and now.microsecond==0:
        current=datetime.now(tz)
        now=now.replace(second=current.second,microsecond=current.microsecond)
    ts_str=now.strftime("%Y-%m-%d %H:%M:%S.%f") + "+00"
    databaze.table("attendance").insert({"user_code":user_code,"position":position,"action":action,"timestamp":ts_str,"valid":True}).execute()
    return True

# --- Excel export functions (nezmenené) ---
from datetime import timedelta as _tdelta
from datetime import time as _time
def get_chip_assignments(df_raw: pd.DataFrame, monday):
    assignments={}
    if df_raw.empty:
        return assignments
    df_raw["timestamp"]=pd.to_datetime(df_raw["timestamp"],errors="coerce")
    df_raw["date"]=df_raw["timestamp"].dt.date
    for pos in df_raw["position"].unique():
        pos_df=df_raw[df_raw["position"]==pos]
        for i in range(7):
            d=monday+_tdelta(days=i)
            day_df=pos_df[pos_df["date"]==d]
            if day_df.empty: continue
            pairs=get_user_pairs(day_df)
            for user,pair in pairs.items():
                if pd.isna(pair["pr"]) or pd.isna(pair["od"]): continue
                pr_t=pair["pr"].time()
                od_t=pair["od"].time()
                if pr_t <= time(7,0) and od_t <= time(15,0):
                    shift="06:00-14_00"
                elif pr_t >= time(13,0) and od_t >= time(21,0):
                    shift="14:00-22:00"
                elif pr_t <= time(7,0) and (od_t >= time(21,0) or od_t < time(2,0)):
                    assignments[(pos,"06:00-14_00",i)]=assignments.get((pos,"06:00-14_00",i),[])+[user]
                    assignments[(pos,"14:00-22_00",i)]=assignments.get((pos,"14:00-22_00",i),[])+[user]
                    continue
                else:
                    continue
                assignments[(pos,shift,i)]=assignments.get((pos,shift,i),[])+[user]
    return assignments

def excel_with_colors(df_matrix, df_day_details, df_raw, monday):
    wb=Workbook()
    ws1=wb.active
    ws1.title="Týždenný prehľad"
    green=PatternFill(start_color="C6EFCE",end_color="C6EFCE",fill_type="solid")
    yellow=PatternFill(start_color="FFEB9C",end_color="FFEB9C",fill_type="solid")
    for r in dataframe_to_rows(df_matrix.reset_index().rename(columns={"index":"Pozícia"}),index=False,header=True):
        ws1.append(r)
    for row in ws1.iter_rows(min_row=2,min_col=2,max_col=1+len(df_matrix.columns),max_row=1+len(df_matrix)):
        for cell in row:
            val=cell.value
            if isinstance(val,(int,float)):
                cell.fill=green
            elif isinstance(val,str) and val.strip().startswith("⚠"):
                cell.fill=yellow
    ws2=wb.create_sheet("Denné - detail")
    for r in dataframe_to_rows(df_day_details,index=False,header=True):
        ws2.append(r)
    ws3=wb.create_sheet("Surové dáta")
    for r in dataframe_to_rows(df_raw,index=False,header=True):
        ws3.append(r)
    ws4=wb.create_sheet("Rozpis čipov")
    days=["pondelok","utorok","streda","štvrtok","piatok","sobota","nedeľa"]
    ws4.append(["position","shift"]+days)
    chip_map=get_chip_assignments(df_raw,monday)
    POS=sorted(df_raw["position"].unique()) if not df_raw.empty else POSITIONS
    for pos in POS:
        for shift in ["06:00-14_00","14:00-22:00"]:
            row_vals=[]
            for i in range(7):
                users=chip_map.get((pos,shift,i),[])
                row_vals.append(", ".join(users) if users else "")
            ws4.append([pos,shift]+row_vals)
    for col in ws4.columns:
        for cell in col:
            cell.alignment=Alignment(horizontal="center",vertical="center")
    out=BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# --- Streamlit UI (nezmenené) ---
st.title("🕓 Admin — Dochádzka (Denný + Týždenný prehľad)")
if "admin_logged" not in st.session_state:
    st.session_state.admin_logged=False
if not st.session_state.admin_logged:
    st.sidebar.header("Admin prihlásenie")
    pw=st.sidebar.text_input("Heslo",type="password")
    if st.sidebar.button("Prihlásiť"):
        if ADMIN_PASS and pw==ADMIN_PASS:
            st.session_state.admin_logged=True
            st.experimental_rerun()
        else:
            st.sidebar.error("Nesprávne heslo alebo ADMIN_PASS nie je nastavené.")
if not st.session_state.admin_logged:
    st.stop()

today=datetime.now(tz).date()
week_ref=st.sidebar.date_input("Vyber deň v týždni (týždeň začína pondelkom):",value=today)
monday=week_ref-timedelta(days=week_ref.weekday())
start_dt=tz.localize(datetime.combine(monday,time(0,0)))
end_dt=tz.localize(datetime.combine(monday+timedelta(days=7),time(0,0)))
df_week=load_attendance(start_dt,end_dt)

default_day=today if monday<=today<=monday+timedelta(days=6) else monday
selected_day=st.sidebar.date_input("Denný prehľad - vyber deň",value=default_day,min_value=monday,max_value=monday+timedelta(days=6))
df_day=df_week[df_week["date"]==selected_day] if not df_week.empty else pd.DataFrame()

if df_week.empty:
    st.warning("Rozsah nie je dostupný v DB (žiadne dáta pre vybraný týždeň).")
else:
    summary=summarize_day(df_day,selected_day)

st.header(f"✅ Denný prehľad — {selected_day.strftime('%A %d.%m.%Y')}")
cols=st.columns(3)
day_details_rows=[]
for i,pos in enumerate(summary.keys()):
    col=cols[i%3]
    info=summary[pos]
    m=info["morning"]
    p=info["afternoon"]
    col.markdown(f"### **{pos}**")
    col.markdown(f"**Ranná:** {m['status']} — {m['hours']} h")
    col.markdown(f"**Poobedná:** {p['status']} — {p['hours']} h")
    if info["details"]:
        for d in info["details"]:
            col.error(d)
    day_details_rows.append({
        "position": pos,
        "morning_status": m['status'],
        "morning_hours": m.get('hours',0),
        "morning_detail": m.get('detail') or "-",
        "afternoon_status": p['status'],
        "afternoon_hours": p.get('hours',0),
        "afternoon_detail": p.get('detail') or "-",
        "total_hours": info.get("total_hours",0)
    })

df_day_details=pd.DataFrame(day_details_rows)

st.header(f"📊 Týždenný prehľad — {monday} – {monday+timedelta(days=6)}")
week_matrix=[]
for pos in POSITIONS:
    row={"position":pos}
    for i in range(7):
        d=monday+timedelta(days=i)
        day_df=df_week[df_week["date"]==d]
        info=summarize_day(day_df,d) if not day_df.empty else {}
        if pos in info:
            row[d.strftime("%a %d.%m")]=info[pos]["total_hours"]
        else:
            row[d.strftime("%a %d.%m")]=0
    week_matrix.append(row)
df_matrix=pd.DataFrame(week_matrix).set_index("position")
st.dataframe(df_matrix,width=1200,height=400)

st.download_button("⬇️ Export Excel",data=excel_with_colors(df_matrix,df_day_details,df_week,monday),file_name=f"dochadzka_{monday}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
